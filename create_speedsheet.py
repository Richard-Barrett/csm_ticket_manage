import json
from ConfigParser import SafeConfigParser
from simple_salesforce import Salesforce
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import dateutil.parser
from pytz import timezone
import pytz
from datetime import datetime, timedelta

### Personal Credentials
parser = SafeConfigParser()
parser.read('salesforce.conf')
sf_url = parser.get('SalesForce', 'url')
sf_usr = parser.get('SalesForce', 'username')
sf_pwd = parser.get('SalesForce', 'password')
sf_tkn = parser.get('SalesForce', 'token')
sf_account = parser.get('SalesForce', 'account')
sf_cloud = parser.get('SalesForce', 'cloud')
last_time = parser.get('SalesForce', 'last_time')
now = parser.get('SalesForce', 'now')
tz  = parser.get('SalesForce', 'tz')
duration = int(parser.get('SalesForce', 'duration'))
report_file_name  = parser.get('SalesForce', 'report_file_name')

### L1 Variables
font_size=10
default_font = Font(name="Open Sans",size=font_size)
defautl_font_hyperlink = Font(name="Open Sans",size=font_size, underline="single", color='4169e1')
defautl_font_bold = Font(name="Open Sans",size=font_size, bold=True)
header_color = '87cefa'
sev_colors = {
   "1":'ff0000',
   "2":'ff6347',
   "3":'ffff00',
   "4":'f5deb3'
}
border = Border(
    left=Side(
        border_style="thin",
        color="000000"
    ),
    right=Side(
        border_style="thin",
        color="000000"
    ),
    top=Side(
        border_style="thin",
        color="000000"
    ),
    bottom=Side(
        border_style="thin",
        color="000000"

    )
)
ticket_fields = [
        "ID",
        "CaseNumber",
        "CreatedById",
        "Severity_Level__c",
        "CreatedDate",
        "Case_record_type__c",
        "RecordTypeId",
        "Resolution_Time_is_violated__c",
        "IsClosed",
        "isMosAlert__c",
        "ClosedDate",
        "Status",
        "Subject",
        "Maintenance_Window_Link__c",
        "MW_Start__c",
        "MW_Actual_End__c",
        "Implemented_Result__c",
        "Risk_Level__c",
        "URL__c",
        "Environment2__c",
        "L2__c",
        "Resolution_Time_DDHHMM__c",
        "SLA_resolution_time__c",
        "Case_Link__c",
        "Customer_Wait_Time_in_Hours__c"
]
query_field = [
        "AccountId = '%s'" % sf_account,
        "Environment2__c = '%s'" % sf_cloud,
        "isMosAlert__c = false",
        "status != 'Canceled'",
        "status != 'Closed'",
        "( ClosedDate = null or (ClosedDate > %s and ClosedDate < %s)) and CreatedDate < %s" % ( last_time, now, now )
]

### Collect data
sf = Salesforce(instance_url=sf_url, username=sf_usr, password=sf_pwd, security_token=sf_tkn)

record_types = sf.query("SELECT ID,NAME from RecordType")['records']
customers = sf.query("SELECT ID,NAME from User where AccountId = '%s'" % sf_account)['records']
cases = sf.query("SELECT %s from Case where %s"
                % ( ','.join(ticket_fields),
                    ' and '.join(query_field) ))['records']
for record_type in record_types:
    if record_type["Name"] == "Technical Case":
        tech_case_id = record_type["Id"]
    if record_type["Name"] == "Change Request":
        change_request_id = record_type["Id"]
    if record_type["Name"] == "Closed Case":
        closed_case_id = record_type["Id"]
customer_ids = []
for customer in customers:
    customer_ids.append(customer['Id'])

### L2 Definition
summary_logic = [
    {
        "name" : "Total Cases",
        "logic" : "COUNT()",
        "result" : "totalSize",
        "duration": 0,
        "base_query" : [
           "AccountId = '%s'" % sf_account,
           "Environment2__c = '%s'" % sf_cloud,
           "isMosAlert__c = false",
           "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
        ],
        "time_query": [
           "CreatedDate < "
        ]
    },
   {
       "name" : "Opened Sev1 Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "Severity_Level__c = 'Sev 1'"
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
       "name" : "Opened Sev2 Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "Severity_Level__c = 'Sev 2'"
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
       "name" : "Opened Sev3 Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "Severity_Level__c = 'Sev 3'"
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
       "name" : "Opened Sev4 Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "Severity_Level__c = 'Sev 4'"
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
       "name" : "Closed Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < "
       ]
   },
   {
       "name" :"Cases Opened By Customer",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
          "( CreatedById = '%s' )" % "' Or CreatedById = '".join(customer_ids)
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
       "name" : "Technical Escalated Cases",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "L2__c = true",
          "( RecordTypeId = '%s' or RecordTypeId = '%s' )" % ( tech_case_id, closed_case_id),
       ],
       "time_query": [
          "CreatedDate < "
       ]
   },
   {
      "name": "%s Days Average resolution time Sev 2 (in hours)" % duration,
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": duration,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 2'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < ",
          "ClosedDate > "
       ]
   },
   {
      "name": "%s Days Average resolution time Sev 3 (in hours)" % duration,
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": duration,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 3'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < ",
          "ClosedDate > "
       ]
   },
   {
      "name": "%s Days Average resolution time Sev 4 (in hours)" % duration,
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": duration,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 4'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < ",
          "ClosedDate > "
       ]
   },
   {
      "name": "All Time Average resolution time Sev 2 (in hours)",
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": 0,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 2'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < "
       ]
   },
   {
      "name": "All Time Average resolution time Sev 3 (in hours)",
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": 0,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 3'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < "
       ]
   },
   {
      "name": "All Time Average resolution time Sev 4 (in hours)",
      "logic" : "AVG(Customer_Wait_Time_in_Hours__c)aver",
      "result" : "aver",
      "duration": 0,
      "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s' " % tech_case_id,
          "Severity_Level__c = 'Sev 4'",
          "ClosedDate != null",
       ],
       "time_query": [
          "ClosedDate < "
       ]
   },
   {
       "name" : "Completed Change Requests",
       "logic" : "COUNT()",
       "result" : "totalSize",
       "duration": 0,
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s'" % change_request_id,
          "ClosedDate != null",
          "status != 'Canceled'"
       ],
       "time_query": [
          "ClosedDate < "
       ]
   },
   {
       "name" : "Opened Change Requests",
       "logic" : "COUNT()",
       "duration": 0,
       "result" : "totalSize",
       "base_query" : [
          "AccountId = '%s'" % sf_account,
          "Environment2__c = '%s'" % sf_cloud,
          "isMosAlert__c = false",
          "RecordTypeId = '%s'" % change_request_id,
          "ClosedDate = null",
          "status != 'Canceled'"
       ],
       "time_query": [
          "CreatedDate < "
       ]
   }

]
summary_page = [
    [
        "Title",
        dateutil.parser.parse(now).strftime('%Y/%m/%d'),
        dateutil.parser.parse(last_time).strftime('%Y/%m/%d'),
        "Delta"
    ]
]
solved_tech_cases = [ 
    [ 
        "Case Number",
        "Severity",
        #"SLA Violated",
        "Resolution time",
        "Created By Customer",
        "Change Request",
        "Created",
        "Closed",
        "Subject"
    ]
]
open_tech_cases = [
    [
        "Case Number",
        "Status",
        "Severity",
        #"SLA Violated",
        "Resolution time",
        "Created By Customer",
        "Escalated",
        "Change Request",
        "Created",
        "Subject"
    ]
]
completed_change_requests = [
    [
        "Case Number",
        "Parent Case",
        "Result",
        "Start",
        "End",
        "Subject"
    ]
]
open_change_requests = [
    [
        "Case Number",
        "Parent Case",
        "Result",
        "Risk",
        "Created",
        "Subject"
    ]
]
#violated_sla_cases = [
#    [
#        "Case Number",
#        "Severity",
#        "Created By Customer",
#        "Escalated",
#        "Created",
#        "Closed",
#        #"Resolution Time",
#        "Subject",
#        "Reason for Late resolution"
#    ]
#]

## Create CSV
for i in range(len(summary_logic)):
    summary=summary_logic[i]["name"]
    row=[summary]
    for date in [ now, last_time ]:
      time_array=[ "%s %s" % (summary_logic[i]["time_query"][0],date) ] 
      if summary_logic[i]["duration"] > 0:
         duration_date=dateutil.parser.parse(date) - timedelta(days=summary_logic[i]["duration"])
         time_array.append("%s %s" % (summary_logic[i]["time_query"][1],
             duration_date.strftime('%Y-%m-%dT%H:%M:%S%z') )) 
      query_string = "SELECT %s from Case where %s and %s" % (
          summary_logic[i]["logic"] ,
          ' and '.join(summary_logic[i]["base_query"]) ,
          ' and '.join(time_array)
          )
      if summary_logic[i]["result"] == "totalSize":
         row.append( sf.query(query_string)['totalSize'])
      else:
         row.append( sf.query(query_string)['records'][0][summary_logic[i]["result"]])
    try:
      row.append(row[1] - row[2])
    except:
      row.append(0)
    summary_page.append(row)

for case in cases:
    for record_type in record_types:
        if record_type["Id"] == case["RecordTypeId"]:
            if record_type["Name"] == "Technical Case" and case['Status'] == 'Solved':
                if any(customer['Id'] == case["CreatedById"] for customer in customers):
                    isCustomerCreated = True
                else:
                    isCustomerCreated = False
                row = [ case['CaseNumber'], case['Severity_Level__c'], #case['Resolution_Time_is_violated__c'], 
                        case['Customer_Wait_Time_in_Hours__c'],
                        isCustomerCreated, case['Id'], case['CreatedDate'], case['ClosedDate'], case['Subject'] ]
                solved_tech_cases.append(row)
                #if case['Resolution_Time_is_violated__c'] == True:
                #    row = [ case['CaseNumber'], case['Severity_Level__c'], isCustomerCreated, case['L2__c'], case['CreatedDate'], case['ClosedDate'], # case['Resolution_Time_DDHHMM__c'],
                #             case['Subject'] , "" ]
                #    violated_sla_cases.append(row)
            if record_type["Name"] == "Technical Case" and case['Status'] != 'Solved':
                if any(customer['Id'] == case["CreatedById"] for customer in customers):
                    isCustomerCreated = True
                else:
                    isCustomerCreated = False
                row = [ case['CaseNumber'], case['Status'], case['Severity_Level__c'], #case['Resolution_Time_is_violated__c'], 
                        case['Customer_Wait_Time_in_Hours__c'],
                        isCustomerCreated, case['L2__c'], case['Id'], case['CreatedDate'], case['Subject'] ]
                open_tech_cases.append(row)
            if record_type["Name"] == "Change Request" and case['Status'] == 'Completed':
                maint_start = case['CreatedDate']
                maint_end = case['ClosedDate']
                for sub_case in cases:
                    if sub_case['Id'] == case['Maintenance_Window_Link__c']:
                        maint_start = sub_case['MW_Start__c']
                        maint_end = sub_case['MW_Actual_End__c']
                        break
                row  = [ case['CaseNumber'], case['Case_Link__c'], case['Implemented_Result__c'], maint_start, maint_end, case['Subject'] ]
                completed_change_requests.append(row)
            if record_type["Name"] == "Change Request" and case['Status'] != 'Completed':
                row  = [ case['CaseNumber'],  case['Case_Link__c'], case['Status'], case['Risk_Level__c'], case['CreatedDate'], case['Subject'] ]
                open_change_requests.append(row)

formated_cases = [
        { 
            "name": "summary_page",
            "csv" : summary_page
        },
        {
            "name": "solved_tech_cases",
            "csv" : solved_tech_cases
        },
        {
            "name": "open_tech_cases",
            "csv" : open_tech_cases
        },
        {
            "name": "completed_change_requests",
            "csv" : completed_change_requests
        },
        {
            "name": "open_change_requests",
            "csv" : open_change_requests
        },
        #{
        #    "name": "violated_sla_cases",
        #    "csv" : violated_sla_cases
        #}
]
## Create Excel
wb = openpyxl.Workbook()
for sheet_num in range(len(formated_cases)):
    case_type = formated_cases[sheet_num]["name"]
    wb.create_sheet(case_type, sheet_num)
    wb.active = sheet_num
    ws = wb.active
    for i in range(len(formated_cases[sheet_num]["csv"])):
        for j in range(len(formated_cases[sheet_num]["csv"][i])):
            try:
                convert_date = dateutil.parser.parse(formated_cases[sheet_num]["csv"][i][j])
                ws.cell(row=i+1, column=j+1).value = convert_date.astimezone(timezone(tz))
            except:
                ws.cell(row=i+1, column=j+1).value=formated_cases[sheet_num]["csv"][i][j]
            ws.cell(row=i+1, column=j+1).border=border
            ws.cell(row=i+1, column=j+1).font=default_font
            if i == 0:
                ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=header_color)
                ws.cell(row=i+1, column=j+1).font= defautl_font_bold
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Severity":
                ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors[formated_cases[sheet_num]["csv"][i][j].split("Sev ")[1]])
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "SLA Violated" and formated_cases[sheet_num]["csv"][i][j] == True:
                ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["1"])
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Title":
                ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=header_color)
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Risk":
                if formated_cases[sheet_num]["csv"][i][j] == "High":
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["1"])
                if formated_cases[sheet_num]["csv"][i][j] == "Medium":
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["2"])
                if formated_cases[sheet_num]["csv"][i][j] == "Low":
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["3"])
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Resolution time":
                if formated_cases[sheet_num]["csv"][i][1] == "Sev 1":
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["1"])
                if formated_cases[sheet_num]["csv"][i][1] == "Sev 2" and formated_cases[sheet_num]["csv"][i][j] > 24:
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["3"])
                if formated_cases[sheet_num]["csv"][i][1] == "Sev 3" and formated_cases[sheet_num]["csv"][i][j] > 120:
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["3"])
                if formated_cases[sheet_num]["csv"][i][1] == "Sev 4" and formated_cases[sheet_num]["csv"][i][j] > 240:
                    ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=sev_colors["3"])
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Change Request":
               try:
                 value=sf.query("SELECT CaseNumber from Case where Environment2__c = '%s' and RecordTypeId = '%s' and Case_Link__c ='%s'"% ( sf_cloud, change_request_id, formated_cases[sheet_num]["csv"][i][j]))['records'][0]['CaseNumber'] 
                 ws.cell(row=i+1, column=j+1).value=value
               except:
                 ws.cell(row=i+1, column=j+1).value=""
            if j == 0 and i != 0:
                for case in cases:
                    if case['CaseNumber'] == formated_cases[sheet_num]["csv"][i][j]:
                        ws.cell(row=i+1, column=j+1).value="=HYPERLINK(\"%s\",\"%s\")" % ( case['URL__c'] , formated_cases[sheet_num]["csv"][i][j] )
                        ws.cell(row=i+1, column=j+1).font= defautl_font_hyperlink
                        break
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "Parent Case":
                value=sf.query("SELECT CaseNumber from Case where Id='%s'"% formated_cases[sheet_num]["csv"][i][j])['records'][0]['CaseNumber']
                ws.cell(row=i+1, column=j+1).value=value
                #for case in cases:
                #    if case['Id'] == formated_cases[sheet_num]["csv"][i][j]:
                #        ws.cell(row=i+1, column=j+1).value=case['CaseNumber']
                #        break
            if type(formated_cases[sheet_num]["csv"][i][j]) is bool:
                if formated_cases[sheet_num]["csv"][i][j] == True:
                    ws.cell(row=i+1, column=j+1).value="Y"
                else:
                    ws.cell(row=i+1, column=j+1).value=""
            if isinstance(formated_cases[sheet_num]["csv"][i][j],float):
                ws.cell(row=i+1, column=j+1).value=round(formated_cases[sheet_num]["csv"][i][j],2)
    
    dims = {}
    i=0
    for row in ws.rows:
        j=0
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(formated_cases[sheet_num]["csv"][i][j]))))
            j+=1
        i+=1
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
ws.active = 0
wb.save(report_file_name)
