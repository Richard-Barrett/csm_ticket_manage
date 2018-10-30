import requests
import json
from ConfigParser import SafeConfigParser
from simple_salesforce import Salesforce
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import dateutil.parser
from pytz import timezone
import pytz
from datetime import datetime, timedelta
import re

### Personal Credentials
parser = SafeConfigParser()
parser.read('salesforce.conf')
sf_url = parser.get('SalesForce', 'url')
sf_usr = parser.get('SalesForce', 'username')
sf_pwd = parser.get('SalesForce', 'password')
sf_tkn = parser.get('SalesForce', 'token')
sf_account = parser.get('SalesForce', 'account')
sf_cloud = parser.get('SalesForce', 'cloud')
last_time = parser.get('SalesForce', 'last_time')
now = parser.get('SalesForce', 'now')
tz  = parser.get('SalesForce', 'tz')
duration = int(parser.get('SalesForce', 'duration'))
report_file_name  = parser.get('SalesForce', 'report_file_name')

parser.read('jira.conf')
jira_url = parser.get('jira', 'url')
jira_usr = parser.get('jira', 'user')
jira_pwd = parser.get('jira', 'password')
headers = {
    'Content-Type': 'application/json',
}


### L1 Variables
font_size=10
default_font = Font(name="Open Sans",size=font_size)
defautl_font_hyperlink = Font(name="Open Sans",size=font_size, underline="single", color='4169e1')
defautl_font_bold = Font(name="Open Sans",size=font_size, bold=True)
header_color = '87cefa'
sev_colors = {
   "1":'ff0000',
   "2":'ff6347',
   "3":'ffff00',
   "4":'f5deb3'
}
border = Border(
    left=Side(
        border_style="thin",
        color="000000"
    ),
    right=Side(
        border_style="thin",
        color="000000"
    ),
    top=Side(
        border_style="thin",
        color="000000"
    ),
    bottom=Side(
        border_style="thin",
        color="000000"

    )
)
ticket_fields = [
        "ID",
        "CaseNumber",
        "Severity_Level__c",
        "CreatedDate",
        "Case_record_type__c",
        "RecordTypeId",
        "Resolution_Time_is_violated__c",
        "IsClosed",
        "isMosAlert__c",
        "ClosedDate",
        "Status",
        "Subject",
        "Closure_Class_Detail__c",
        "Launch_Pad_URL_1__c",
        "Launch_Pad_URL_2__c",
        "Launch_Pad_URL_3__c",
        "Launch_Pad_URL_4__c",
]
query_field = [
        "AccountId = '%s'" % sf_account,
        "Environment2__c = '%s'" % sf_cloud,
        "( Launch_Pad_URL_1__c != null or Launch_Pad_URL_2__c != null or Launch_Pad_URL_3__c != null or Launch_Pad_URL_4__c != null )",
        "( ClosedDate = null or Closure_Class_Detail__c = 'Coding / Missing Functionality (New Blueprint Filed)' or Closure_Class_Detail__c = 'Coding / Errant Functionality (LP Bug Found)' )"
]

### Collect data
sf = Salesforce(instance_url=sf_url, username=sf_usr, password=sf_pwd, security_token=sf_tkn)
cases = sf.query("SELECT %s from Case where %s"
                % ( ','.join(ticket_fields),
                    ' and '.join(query_field) ))['records']

bug_fixes = [
    [
        "Case Number",
        "Subject",
        "Status",
        "Closed Reason",
        "PROD-Ticket ID",
        "PROD-Ticket Status",
        "PROD-Ticket Subject",
        "PROD-Ticket Fixed Version"
    ]
]

for case in cases:
    if case['Launch_Pad_URL_1__c'] != None:
        row = [ case['CaseNumber'], case['Subject'], case['Status'], case['Closure_Class_Detail__c'], case['Launch_Pad_URL_1__c'], "", "","" ]
        bug_fixes.append(row)
    if case['Launch_Pad_URL_2__c'] != None:
        row = [ case['CaseNumber'], case['Subject'], case['Status'], case['Closure_Class_Detail__c'], case['Launch_Pad_URL_2__c'], "", "","" ]
        bug_fixes.append(row)
    if case['Launch_Pad_URL_3__c'] != None:
        row = [ case['CaseNumber'], case['Subject'], case['Status'], case['Closure_Class_Detail__c'], case['Launch_Pad_URL_3__c'], "", "","" ]
        bug_fixes.append(row)
    if case['Launch_Pad_URL_4__c'] != None:
        row = [ case['CaseNumber'], case['Subject'], case['Status'], case['Closure_Class_Detail__c'], case['Launch_Pad_URL_4__c'], "", "","" ]
        bug_fixes.append(row)
formated_cases = [
        { 
            "name": "bug_fixes",
            "csv" : bug_fixes
        }
]
## Create Excel
wb = openpyxl.load_workbook(report_file_name)
last_sheet=-1
for sheet in wb:
    last_sheet+=1
for sheet_num in range(len(formated_cases)):
    case_type = formated_cases[sheet_num]["name"]
    wb.create_sheet(case_type, last_sheet + sheet_num)
    wb.active = last_sheet + sheet_num
    ws = wb.active
    for i in range(len(formated_cases[sheet_num]["csv"])):
        for j in range(len(formated_cases[sheet_num]["csv"][i])):
            try:
                convert_date = dateutil.parser.parse(formated_cases[sheet_num]["csv"][i][j])
                ws.cell(row=i+1, column=j+1).value = convert_date.astimezone(timezone(tz))
            except:
                ws.cell(row=i+1, column=j+1).value=formated_cases[sheet_num]["csv"][i][j]
            ws.cell(row=i+1, column=j+1).border=border
            ws.cell(row=i+1, column=j+1).font=default_font
            if i == 0:
                ws.cell(row=i+1, column=j+1).fill = PatternFill(fill_type='solid', fgColor=header_color)
                ws.cell(row=i+1, column=j+1).font= defautl_font_bold
            if i != 0 and formated_cases[sheet_num]["csv"][0][j] == "PROD-Ticket ID":
                r = re.compile("^.*/([A-Z]*-[0-9]*).*$")
                prod_id=r.search(formated_cases[sheet_num]["csv"][i][j]).group(1)
                ws.cell(row=i+1, column=j+1).value=prod_id

                response = requests.get("%s/rest/api/2/issue/%s" % (jira_url, prod_id) , headers=headers, auth=(jira_usr, jira_pwd))

                formated_cases[sheet_num]["csv"][i][j+1]=response.json()['fields']['status']['name']
                formated_cases[sheet_num]["csv"][i][j+2]=response.json()['fields']['summary']
                formated_cases[sheet_num]["csv"][i][j+3]=""
                for version in response.json()['fields']['fixVersions']:
                  formated_cases[sheet_num]["csv"][i][j+3]+=version['name']
            
            if type(formated_cases[sheet_num]["csv"][i][j]) is bool:
                if formated_cases[sheet_num]["csv"][i][j] == True:
                    ws.cell(row=i+1, column=j+1).value="Y"
                else:
                    ws.cell(row=i+1, column=j+1).value=""
            if isinstance(formated_cases[sheet_num]["csv"][i][j],float):
                ws.cell(row=i+1, column=j+1).value=round(formated_cases[sheet_num]["csv"][i][j],2)
    
    dims = {}
    i=0
    for row in ws.rows:
        j=0
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
            j+=1
        i+=1
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
wb.active = 0
wb.save(report_file_name)
